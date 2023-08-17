from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter 

wb = load_workbook('tim.xlsx')
ws = wb.active

ws.merge_cells("A1:D1")
ws.unmerge_cells("A1:D1")

wb.save("tim.xlsx")