from openpyxl import Workbook, load_workbook

wb = load_workbook('excel with python.xlsx')
ws = wb.active

print(ws['A2'].value)
ws['A2'].value = 'deneme'
print(ws['A2'].value)
ws['A3'].value = 'zazazasdasd'
print(ws['A3'].value)


wb.save('excel with python.xlsx')