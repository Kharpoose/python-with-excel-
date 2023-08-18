from openpyxl import Workbook, load_workbook

wb = load_workbook('excelwithpython.xlsx')
ws = wb.active

print(ws['A2'].value)
ws['A2'].value = 'deneme'
print(ws['A2'].value)
print(wb.sheetnames)

wb.create_sheet('Test1') #Create  new sheet wow

wb.save('excelwithpython.xlsx')